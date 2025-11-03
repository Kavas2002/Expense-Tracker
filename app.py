import streamlit as st
import pandas as pd
from datetime import date
import os
from openpyxl import Workbook, load_workbook

st.set_page_config(page_title="Expense Entry Form")
st.title("💰 CLOUD EXPENSE TRACKER")

# File setup (Excel, not CSV)
excel_file = "expenses.xlsx"

# Create Excel if not exists
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Expense Type", "Category", "Description", "Amount", "Payment Mode", "Location", "Receipt", "Notes"])
    wb.save(excel_file)

# Expense Entry Form
with st.form("expense_form"):
    st.subheader("Enter Expense Details")

    expense_date = st.date_input("Date", date.today())
    expense_type = st.selectbox("Expense Type", ["Personal", "Business", "Friends", "Relative", "Debt"])
    category = st.selectbox("Category", ["Food", "Transport", "Shopping", "Bills", "Other"])
    description = st.text_input("Description (Product/Service Purchased)")
    amount = st.number_input("Amount (₹)", min_value=0.0, format="%.2f")
    payment_mode = st.selectbox("Payment Mode", ["Cash", "UPI", "Card", "Other"])
    location = st.text_input("Location")
    receipt = st.file_uploader("Upload Receipt (Optional)", type=["jpg", "jpeg", "png", "pdf"])
    notes = st.text_input("Notes")
    confirm = st.checkbox("✅ Confirm to add this expense")

    submitted = st.form_submit_button("Add Expense")

    if submitted:
        if not confirm:
            st.warning("⚠️ Please confirm before adding the expense.")
        elif amount <= 0:
            st.warning("⚠️ Please enter a valid amount.")
        elif description.strip() == "":
            st.warning("⚠️ Please enter a description.")
        elif location.strip() == "":
            st.warning("⚠️ Please enter the location.")
        else:
            # Save receipt if uploaded
            receipt_path = None
            if receipt is not None:
                receipts_folder = "receipts"
                os.makedirs(receipts_folder, exist_ok=True)
                receipt_path = os.path.join(receipts_folder, receipt.name)
                with open(receipt_path, "wb") as f:
                    f.write(receipt.getbuffer())

            # Append data to Excel file
            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([
                str(expense_date),
                expense_type,
                category,
                description,
                amount,
                payment_mode,
                location,
                receipt_path if receipt_path else "",
                notes
            ])
            wb.save(excel_file)

            st.success("✅ Expense saved successfully to Excel!")

